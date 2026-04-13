# -*- coding: utf-8 -*-
"""
抽出メインロジックモジュール
"""
import os
import math
import time
import traceback
import re
import datetime
import openpyxl
from openpyxl.styles import numbers
from dxf_core import get_all_elements_from_dxf, apply_text_inheritance, zen_to_han

def _convert_value_for_excel(val_str, format_type):
    if not val_str:
        return val_str
    
    val_str = val_str.strip()
    if format_type == "文字列":
        return str(val_str)
        
    if format_type in ["数値", "通貨", "会計", "分数", "指数"]:
        try:
            clean_str = val_str.replace(",", "").replace("¥", "").replace("\\", "").replace("円", "").strip()
            if "." in clean_str: return float(clean_str)
            return int(clean_str)
        except ValueError:
            return val_str
            
    if format_type == "パーセンテージ":
        try:
            clean_str = val_str.replace(",", "").replace("%", "").strip()
            return float(clean_str) / 100.0 if "%" not in val_str else float(clean_str)
        except ValueError:
            return val_str
            
    if format_type == "日付":
        nums = re.findall(r'\d+', val_str)
        if len(nums) >= 3:
            try:
                y, m, d = int(nums[0]), int(nums[1]), int(nums[2])
                if y < 100: y += 2000
                return datetime.date(y, m, d)
            except ValueError:
                pass
        return val_str

    if format_type == "時刻":
        nums = re.findall(r'\d+', val_str)
        if len(nums) >= 2:
            try:
                h, m = int(nums[0]), int(nums[1])
                s = int(nums[2]) if len(nums) >= 3 else 0
                return datetime.time(h, m, s)
            except ValueError:
                pass
        return val_str
            
    return val_str

def run_extract_dxf(target_files, save_dir, is_keyword_mode, y_threshold, base_kw_str, base_kw2_str, base_dist, keyword_settings, progress_callback=None, cancel_check=None):
    try:
        processed_count = 0
        error_logs = []
        total_files = len(target_files)

        if is_keyword_mode:
            if not base_kw_str:
                return False, "第1基準文字が設定されていません。"
            if not keyword_settings:
                return False, "抽出設定（抽出範囲）がありません。「＋ プレビューを開いて基準文字と抽出範囲を設定」から設定を行ってください。"

            base_kw_str = zen_to_han(base_kw_str)
            base_kw2_str = zen_to_han(base_kw2_str)

            out_wb = openpyxl.Workbook()
            ws = out_wb.active
            ws.title = "抽出結果"
            
            ws_coord = out_wb.create_sheet(title="座標リスト(設定確認用)")
            ws_coord.append(["元ファイル名", "テキスト", "X座標", "Y座標"])
            ws_coord.column_dimensions['A'].width = 25
            ws_coord.column_dimensions['B'].width = 40
            ws_coord.column_dimensions['C'].width = 15
            ws_coord.column_dimensions['D'].width = 15

            header = ["元ファイル名"]
            for cfg in keyword_settings:
                header.append(cfg["col_name"])

            all_rows = [header]

            for i, file_path in enumerate(target_files):
                if cancel_check and cancel_check():
                    return False, "ユーザーにより処理が中止されました。"

                if progress_callback:
                    progress_callback(i, total_files, f"抽出中 ({i+1}/{total_files}): {os.path.basename(file_path)}")
                
                texts, _ = get_all_elements_from_dxf(file_path)
                if not texts:
                    error_logs.append(f"【{os.path.basename(file_path)}】テキストが見つかりません。")
                    continue

                for t in texts:
                    ws_coord.append([os.path.basename(file_path), t['text'], round(t['x'], 3), round(t['y'], 3)])

                file_row = [os.path.basename(file_path)]

                kw_candidates = [t for t in texts if base_kw_str == t['text'].replace(" ", "").replace("　", "").lower()]
                if not kw_candidates:
                    kw_candidates = [t for t in texts if base_kw_str in t['text'].replace(" ", "").replace("　", "").lower()]
                    
                kw2_candidates = []
                if base_kw2_str:
                    kw2_candidates = [t for t in texts if base_kw2_str == t['text'].replace(" ", "").replace("　", "").lower()]
                    if not kw2_candidates:
                        kw2_candidates = [t for t in texts if base_kw2_str in t['text'].replace(" ", "").replace("　", "").lower()]
                else:
                    kw2_candidates = [None]
                
                transform_params = None
                found_anchor1 = None
                found_anchor2 = None
                
                for kw_entity in kw_candidates:
                    for kw2_entity in kw2_candidates:
                        ax, ay = kw_entity['x'], kw_entity['y']
                        a2x, a2y = None, None
                        
                        if kw2_entity:
                            a2x, a2y = kw2_entity['x'], kw2_entity['y']
                            if ax == a2x and ay == a2y: continue
                            
                        L, theta = 1.0, 0.0
                        cos_t = 1.0
                        sin_t = 0.0
                        
                        transform_params = (ax, ay, L, cos_t, sin_t)
                        found_anchor1 = kw_entity
                        found_anchor2 = kw2_entity
                        break
                    if transform_params: break
                
                if transform_params:
                    ax, ay, L, cos_t, sin_t = transform_params
                    
                    if base_dist > 0 and found_anchor2:
                        a2x, a2y = found_anchor2['x'], found_anchor2['y']
                        target_dist = math.sqrt((a2x - ax)**2 + (a2y - ay)**2)
                        if target_dist > 0:
                            L = target_dist / base_dist
                    
                    for cfg in keyword_settings:
                        x_min, x_max = cfg["xmin"], cfg["xmax"]
                        y_min, y_max = cfg["ymin"], cfg["ymax"]
                        excludes = [x.strip() for x in cfg.get("exclude", "").split(",") if x.strip()]
                            
                        matched_texts = []
                        for t in texts:
                            if t == found_anchor1 or t == found_anchor2: continue
                            
                            t_clean_str = t['text'].strip()
                            if t_clean_str == found_anchor1['text'].strip() or (found_anchor2 and t_clean_str == found_anchor2['text'].strip()): continue
                            
                            rep_x = t['x']
                            rep_y = t['y'] + (t.get('h', 2.5) * 0.5)
                            
                            px, py = rep_x - ax, rep_y - ay
                            nx = (px * cos_t - py * sin_t) / L
                            ny = (px * sin_t + py * cos_t) / L
                            
                            if x_min <= nx <= x_max and y_min <= ny <= y_max:
                                matched_texts.append(t)
                        
                        if matched_texts:
                            matched_texts.sort(key=lambda item: (-item['y'], item['x']))
                            found_val = "/".join([t['text'] for t in matched_texts])
                            
                            replaces = cfg.get("replaces", [])
                            for rep in replaces:
                                rep_before = rep.get("before", "")
                                rep_after = rep.get("after", "")
                                if rep_before:
                                    try:
                                        found_val = re.sub(rep_before, rep_after, found_val)
                                    except re.error:
                                        found_val = found_val.replace(rep_before, rep_after)
                                    
                            for ex in excludes:
                                try:
                                    found_val = re.sub(ex, "", found_val)
                                except re.error:
                                    found_val = found_val.replace(ex, "")
                        else:
                            found_val = ""
                            
                        found_val = _convert_value_for_excel(found_val, cfg.get("format", "標準"))
                        file_row.append(found_val)
                else:
                    for _ in keyword_settings:
                        file_row.append("")
                    error_logs.append(f"【{os.path.basename(file_path)}】基準文字が見つかりません。")
                
                all_rows.append(file_row)
                processed_count += 1
                
            # 集約データを適用
            apply_text_inheritance(all_rows)
            
            # シートに書き出し
            for row_idx, row_data in enumerate(all_rows):
                ws.append(row_data)
                
                if row_idx > 0:
                    current_row_idx = ws.max_row
                    for col_idx, cfg in enumerate(keyword_settings):
                        cell = ws.cell(row=current_row_idx, column=col_idx + 2)
                        fmt = cfg.get("format", "標準")
                        
                        if fmt == "数値": cell.number_format = '0_ '
                        elif fmt == "通貨": cell.number_format = '"¥"#,##0;[Red]\-"¥"#,##0'
                        elif fmt == "会計": cell.number_format = '_ *"¥"* #,##0_ ;_ *"¥"* \-#,##0_ ;_ *"¥"* "-"_ ;_ @_ '
                        elif fmt == "日付": cell.number_format = 'yyyy/m/d'
                        elif fmt == "時刻": cell.number_format = 'h:mm:ss'
                        elif fmt == "パーセンテージ": cell.number_format = '0.0%'
                        elif fmt == "分数": cell.number_format = '# ?/?'
                        elif fmt == "指数": cell.number_format = '0.00E+00'
                        elif fmt == "文字列": cell.number_format = '@'
                        else: cell.number_format = 'General'

            for col in ws.columns:
                try:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            length = sum(2 if ord(c) > 255 else 1 for c in str(cell.value))
                            if length > max_length: max_length = length
                    ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
                except: pass

            if progress_callback:
                progress_callback(total_files, total_files, "完了しました")
                
            output_name = f"抽出結果集約_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
            if len(target_files) == 1:
                output_name = f"{os.path.splitext(os.path.basename(target_files[0]))[0]}_抽出.xlsx"
            output_path = os.path.join(save_dir, output_name)
            
            try:
                out_wb.save(output_path)
            except PermissionError:
                return False, f"Excelファイルが開かれているため保存できません。\nパス: {output_path}"
            except Exception as e:
                return False, f"保存エラー: {e}"
                
            msg = f"{processed_count} 個のファイルから抽出を完了しました。\n保存先: {output_path}"
            if error_logs: msg += "\n\n※一部エラーあり:\n" + "\n".join(error_logs[:5])
            return True, msg

        else:
            # モードB: 全体抽出モード
            out_wb = openpyxl.Workbook()
            ws = out_wb.active
            ws.title = "抽出データ"
            
            ws_coord = out_wb.create_sheet(title="座標リスト(設定確認用)")
            ws_coord.append(["元ファイル名", "テキスト", "X座標", "Y座標"])
            ws_coord.column_dimensions['A'].width = 25
            ws_coord.column_dimensions['B'].width = 40
            ws_coord.column_dimensions['C'].width = 15
            ws_coord.column_dimensions['D'].width = 15

            for i, file_path in enumerate(target_files):
                if cancel_check and cancel_check():
                    return False, "ユーザーにより処理が中止されました。"

                if progress_callback:
                    progress_callback(i, total_files, f"全体抽出中 ({i+1}/{total_files}): {os.path.basename(file_path)}")
                
                fname = os.path.basename(file_path)
                texts, _ = get_all_elements_from_dxf(file_path)
                if not texts:
                    error_logs.append(f"【{fname}】読込エラー。")
                    continue

                for t in texts:
                    ws_coord.append([fname, t['text'], round(t['x'], 3), round(t['y'], 3)])

                texts.sort(key=lambda item: (-item['y'], item['x']))
                rows, current_row, current_y = [], None, None

                for item in texts:
                    if current_y is None:
                        current_y = item['y']
                        current_row = [item]
                    elif abs(current_y - item['y']) <= y_threshold:
                        current_row.append(item)
                    else:
                        current_row.sort(key=lambda i: i['x'])
                        rows.append([fname] + [i['text'] for i in current_row])
                        current_y = item['y']
                        current_row = [item]
                if current_row:
                    current_row.sort(key=lambda i: i['x'])
                    rows.append([fname] + [i['text'] for i in current_row])

                for row_data in rows: ws.append(row_data)
                processed_count += 1

            for col in ws.columns:
                try:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            length = sum(2 if ord(c) > 255 else 1 for c in str(cell.value))
                            if length > max_length: max_length = length
                    ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
                except: pass

            if progress_callback:
                progress_callback(total_files, total_files, "完了しました")
                
            output_name = f"全体テキスト抽出集約_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
            if len(target_files) == 1:
                output_name = f"{os.path.splitext(os.path.basename(target_files[0]))[0]}_全体テキスト抽出.xlsx"
            output_path = os.path.join(save_dir, output_name)
            
            try:
                out_wb.save(output_path)
            except PermissionError:
                return False, f"Excelが開かれています。\nパス: {output_path}"
            except Exception as e:
                return False, f"エラー: {e}"
                
            msg = f"{processed_count} 個のDXFファイルからの全体抽出が完了しました。\n保存先: {output_path}"
            if error_logs: msg += "\n\n※一部エラーあり:\n" + "\n".join(error_logs[:5])
            return True, msg

    except Exception as e:
        return False, f"予期せぬエラーが発生しました。\n\n【詳細】\n{traceback.format_exc()}"