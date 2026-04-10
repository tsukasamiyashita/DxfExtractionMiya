# -*- coding: utf-8 -*-
"""
DxfExtractionMiya v1.0.0
機能：DXFファイルからのプレビュー＆テキスト抽出（マウスで範囲指定・連続追加）、設定保存、複数Excelの集約
1ファイル完結版
"""

import os
import re
import traceback
import json
import ezdxf
from ezdxf import recover
import openpyxl
from tkinter import *
from tkinter import filedialog, messagebox
from tkinter import ttk

# ==========================
# 共通変数
# ==========================
APP_TITLE = "DxfExtractionMiya"
VERSION = "v1.0.0"

selected_files = []
selected_folder = ""
current_mode = None  # "file" or "folder"

# ==========================
# 選択処理
# ==========================
def select_files():
    global selected_files, selected_folder, current_mode
    files = filedialog.askopenfilenames(
        title="ファイルを選択",
        filetypes=[("すべての対応ファイル", "*.dxf;*.xlsx;*.xlsm;*.xls"), ("DXFファイル", "*.dxf"), ("Excelファイル", "*.xlsx;*.xlsm;*.xls")]
    )
    if files:
        selected_files = list(files)
        selected_folder = ""
        current_mode = "file"
        update_path_display()
        update_button_state()

def select_folder():
    global selected_folder, selected_files, current_mode
    folder = filedialog.askdirectory(title="フォルダを選択")
    if folder:
        selected_folder = folder
        selected_files = []
        current_mode = "folder"
        update_path_display()
        update_button_state()

def update_path_display():
    text_paths.delete(1.0, END)
    if current_mode == "file":
        text_paths.insert(END, "\n".join(selected_files))
    elif current_mode == "folder":
        text_paths.insert(END, f"フォルダ: {selected_folder}")

def update_button_state():
    state = NORMAL if current_mode else DISABLED
    btn_extract.config(state=state)
    btn_aggregate.config(state=state)

# ==========================
# 設定の保存と読み込み
# ==========================
def save_settings():
    try:
        user_dir = os.path.expanduser('~')
        app_dir = os.path.join(user_dir, 'DxfExtractionMiya')
        os.makedirs(app_dir, exist_ok=True)
        
        settings_file = filedialog.asksaveasfilename(
            title="現在の抽出設定を保存",
            initialdir=app_dir,
            initialfile="settings.json",
            defaultextension=".json",
            filetypes=[("JSONファイル", "*.json"), ("すべてのファイル", "*.*")]
        )
        
        if not settings_file:
            return  # キャンセル時
        
        kw_list = []
        for row in keyword_entries:
            kw_list.append({
                "kw": row["kw_var"].get(),
                "anchor": row["anchor_var"].get(),
                "xmin": row["xmin_var"].get(),
                "xmax": row["xmax_var"].get(),
                "ymin": row["ymin_var"].get(),
                "ymax": row["ymax_var"].get(),
                "sample": row.get("sample_text", "")
            })

        settings = {
            "mode": mode_var.get(),
            "save_option": save_option.get(),
            "threshold": threshold_var.get(),
            "keywords": kw_list
        }
        
        with open(settings_file, 'w', encoding='utf-8') as f:
            json.dump(settings, f, ensure_ascii=False, indent=2)
            
        messagebox.showinfo("保存完了", f"設定を保存しました。\n保存先: {settings_file}")
    except Exception as e:
        messagebox.showerror("保存エラー", f"設定の保存中にエラーが発生しました。\n{e}")

def load_settings():
    try:
        user_dir = os.path.expanduser('~')
        app_dir = os.path.join(user_dir, 'DxfExtractionMiya')
        
        settings_file = filedialog.askopenfilename(
            title="保存した抽出設定を読込",
            initialdir=app_dir,
            filetypes=[("JSONファイル", "*.json"), ("すべてのファイル", "*.*")]
        )
        
        if not settings_file:
            return  # キャンセル時
            
        with open(settings_file, 'r', encoding='utf-8') as f:
            settings = json.load(f)
            
        if "mode" in settings:
            mode_var.set(settings["mode"])
            toggle_mode()
        if "save_option" in settings:
            save_option.set(settings["save_option"])
        if "threshold" in settings:
            threshold_var.set(settings["threshold"])
            
        if "keywords" in settings:
            # 現在のUIをクリア
            for row in list(keyword_ui_frames):
                row.destroy()
            keyword_entries.clear()
            keyword_ui_frames.clear()
            
            # 保存データから復元
            for kw_data in settings["keywords"]:
                add_keyword_row(
                    kw=kw_data.get("kw", ""),
                    anchor_val=kw_data.get("anchor", "左下"),
                    xmin=kw_data.get("xmin", 0.0),
                    xmax=kw_data.get("xmax", 0.0),
                    ymin=kw_data.get("ymin", 0.0),
                    ymax=kw_data.get("ymax", 0.0),
                    sample_text=kw_data.get("sample", "")
                )
        
        root.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))
        messagebox.showinfo("読込完了", "設定を読み込みました。")
    except Exception as e:
        messagebox.showerror("読込エラー", f"設定の読込中にエラーが発生しました。\n{e}")

# ==========================
# 共通ユーティリティ（抽出基盤）
# ==========================
def get_save_dir(original_path=None):
    if save_option.get() == 1 and original_path:
        return os.path.dirname(original_path)
    else:
        folder = filedialog.askdirectory(title="保存先フォルダを選択")
        return folder

def sanitize_text(text):
    if text is None: return ""
    text_str = str(text)
    illegal_chars = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    return illegal_chars.sub('', text_str).strip()

def get_point(pt):
    try:
        if pt is None: return 0.0, 0.0
        if hasattr(pt, 'x') and hasattr(pt, 'y'):
            return float(pt.x), float(pt.y)
        elif hasattr(pt, '__getitem__') and len(pt) >= 2:
            return float(pt[0]), float(pt[1])
    except: pass
    return 0.0, 0.0

def get_text_dimensions(entity, text_val):
    height = 2.5
    try:
        if hasattr(entity, 'dxf') and hasattr(entity, 'dxftype') and callable(entity.dxftype):
            if entity.dxftype() == 'MTEXT' and hasattr(entity.dxf, 'char_height'):
                height = float(entity.dxf.char_height)
            elif hasattr(entity.dxf, 'height'):
                height = float(entity.dxf.height)
    except: pass
    width = sum((height * 1.0) if ord(char) > 255 else (height * 0.6) for char in str(text_val))
    return height, width

def extract_text_from_entity(entity):
    text_val = ""
    try:
        if hasattr(entity, 'dxftype') and callable(entity.dxftype):
            if entity.dxftype() == 'MTEXT' and hasattr(entity, 'plain_text'):
                text_val = entity.plain_text()
            elif hasattr(entity, 'dxf') and hasattr(entity.dxf, 'text'):
                text_val = entity.dxf.text
    except: pass

    x, y = 0.0, 0.0
    try:
        if hasattr(entity, 'dxf') and hasattr(entity.dxf, 'insert'):
            x, y = get_point(entity.dxf.insert)
    except: pass
    return text_val, x, y

def get_all_elements_from_dxf(file_path):
    try:
        doc = ezdxf.readfile(file_path)
    except ezdxf.DXFStructureError:
        try: doc, _ = recover.readfile(file_path)
        except: return [], []
    except: return [], []
    
    if doc is None: return [], []
    texts, seen_texts = [], set()
    shapes = []
    
    layouts = []
    try: layouts.append(doc.modelspace())
    except: pass
    try: layouts.extend(list(doc.layouts()))
    except: pass
    
    def process_entity(entity):
        try:
            if entity is None or not hasattr(entity, 'dxftype') or not callable(entity.dxftype): return
            etype = entity.dxftype()
            
            if etype in {'TEXT', 'MTEXT'}:
                text_val, x, y = extract_text_from_entity(entity)
                clean_text = sanitize_text(text_val)
                if clean_text:
                    key = f"{clean_text}_{round(x,2)}_{round(y,2)}"
                    if key not in seen_texts:
                        h, w = get_text_dimensions(entity, clean_text)
                        texts.append({"text": clean_text, "x": x, "y": y, "w": w, "h": h})
                        seen_texts.add(key)
            elif etype == 'LINE':
                shapes.append({'type': 'line', 'start': get_point(entity.dxf.start), 'end': get_point(entity.dxf.end)})
            elif etype in {'LWPOLYLINE', 'POLYLINE'}:
                points = []
                if etype == 'LWPOLYLINE':
                    for pt in entity: points.append(get_point(pt))
                else:
                    for vertex in entity.vertices: points.append(get_point(vertex.dxf.location))
                if points:
                    shapes.append({'type': 'polyline', 'points': points, 'closed': getattr(entity, 'is_closed', False)})
            elif etype == 'CIRCLE':
                shapes.append({'type': 'circle', 'center': get_point(entity.dxf.center), 'radius': float(entity.dxf.radius)})
            elif etype == 'ARC':
                shapes.append({
                    'type': 'arc', 'center': get_point(entity.dxf.center), 'radius': float(entity.dxf.radius),
                    'start_angle': float(entity.dxf.start_angle), 'end_angle': float(entity.dxf.end_angle)
                })
        except: pass

    for layout in layouts:
        try:
            for entity in layout:
                process_entity(entity)
                try:
                    if entity.dxftype() == 'INSERT':
                        if hasattr(entity, 'attribs') and entity.attribs:
                            for attrib in entity.attribs: process_entity(attrib)
                        if hasattr(entity, 'virtual_entities') and callable(entity.virtual_entities):
                            for v_entity in entity.virtual_entities(): process_entity(v_entity)
                except: pass
        except: pass
        
    return texts, shapes

def apply_text_inheritance(final_aggregated_data):
    if len(final_aggregated_data) <= 1: return
    def is_text_to_inherit(text):
        s = str(text).strip()
        if not s or s in ["〃", "”", "\"", "''", "””", "''", "同上", "...", "…"]: return False
        return bool(re.search(r'[a-zA-Zａ-ｚＡ-Ｚぁ-んァ-ン一-龥0-9０-９]', s))
    
    header = final_aggregated_data[0]
    skip_cols = {idx for idx, h in enumerate(header) if "備考" in str(h)}
    
    for col_idx in range(1, len(header)):
        if col_idx in skip_cols: continue
        last_text = ""
        for row_idx in range(1, len(final_aggregated_data)):
            cell_val = str(final_aggregated_data[row_idx][col_idx]).strip()
            if cell_val == "None":
                cell_val, final_aggregated_data[row_idx][col_idx] = "", ""
            if cell_val in ["〃", "”", "\"", "''", "””", "''", "同上", "...", "…"]:
                if last_text: final_aggregated_data[row_idx][col_idx] = last_text
            elif cell_val != "":
                last_text = cell_val if is_text_to_inherit(cell_val) else ""

# ==========================
# プレビュー・範囲選択ダイアログ
# ==========================
class PreviewDialog(Toplevel):
    def __init__(self, parent, dxf_path, on_complete):
        super().__init__(parent)
        self.title(f"プレビュー＆範囲選択 - {os.path.basename(dxf_path)}")
        self.geometry("1200x800")
        
        try: self.state('zoomed')
        except:
            try: self.attributes('-zoomed', True)
            except: pass

        self.transient(parent)
        self.grab_set()
        
        self.on_complete = on_complete
        self.texts, self.shapes = get_all_elements_from_dxf(dxf_path)
        
        self.scale = 1.0
        self.base_scale = 1.0
        self.offset_x = 0.0
        self.offset_y = 0.0
        self.min_x = 0.0
        self.max_y = 0.0
        
        self.mode = StringVar(value="anchor")
        self.anchor = None
        self.rect_dxf_start = None
        self.rect_dxf_end = None
        self.is_dragging = False
        self.pan_start_x = 0
        self.pan_start_y = 0
        
        self.setup_ui()
        self.init_transform()
        self.draw()
        
    def setup_ui(self):
        toolbar = Frame(self, bg="#E9ECEF", pady=10, padx=10)
        toolbar.pack(fill=X)
        
        Label(toolbar, text="手順: ", font=("Meiryo UI", 10, "bold"), bg="#E9ECEF").pack(side=LEFT)
        self.rb1 = Radiobutton(toolbar, text="1. 基準文字を左クリック", variable=self.mode, value="anchor", indicatoron=0, bg="#FFF3CD", selectcolor="#FFC107", padx=10, pady=5)
        self.rb1.pack(side=LEFT, padx=5)
        self.rb2 = Radiobutton(toolbar, text="2. 抽出範囲を左ドラッグ", variable=self.mode, value="rect", indicatoron=0, bg="#D1E7DD", selectcolor="#198754", padx=10, pady=5)
        self.rb2.pack(side=LEFT, padx=5)
        
        Label(toolbar, text="※ ホイール:ズーム  |  右ドラッグ:移動", fg="#6C757D", bg="#E9ECEF").pack(side=LEFT, padx=20)
        
        btn_frame = Frame(toolbar, bg="#E9ECEF")
        btn_frame.pack(side=RIGHT, padx=5)
        Button(btn_frame, text="＋ 設定を追加して次へ", command=self.confirm, bg="#0D6EFD", fg="white", font=("Meiryo UI", 10, "bold"), padx=10).pack(side=LEFT, padx=5)
        Button(btn_frame, text="完了して閉じる", command=self.destroy, bg="#6C757D", fg="white", font=("Meiryo UI", 10, "bold"), padx=10).pack(side=LEFT, padx=5)
        
        # --- リアルタイム抽出プレビュー表示バー ---
        preview_bar = Frame(self, bg="#D1E7DD", pady=8, padx=15)
        preview_bar.pack(fill=X)
        self.preview_label = Label(preview_bar, text="【抽出プレビュー】 1. まずは図面内の基準にする文字をクリックしてください", font=("Meiryo UI", 12, "bold"), bg="#D1E7DD", fg="#0F5132")
        self.preview_label.pack(side=LEFT)
        
        self.canvas = Canvas(self, bg="white", cursor="crosshair")
        self.canvas.pack(fill=BOTH, expand=True)
        
        self.canvas.bind("<ButtonPress-1>", self.on_left_press)
        self.canvas.bind("<B1-Motion>", self.on_left_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_left_release)
        self.canvas.bind("<ButtonPress-3>", self.on_right_press)
        self.canvas.bind("<B3-Motion>", self.on_right_drag)
        self.canvas.bind("<MouseWheel>", self.on_mousewheel)
        
    def init_transform(self):
        xs, ys = [], []
        for t in self.texts:
            xs.append(t["x"])
            ys.append(t["y"])
        for s in self.shapes:
            if s['type'] == 'line':
                xs.extend([s['start'][0], s['end'][0]])
                ys.extend([s['start'][1], s['end'][1]])
            elif s['type'] == 'polyline':
                for pt in s['points']:
                    xs.append(pt[0]); ys.append(pt[1])
            elif s['type'] in {'circle', 'arc'}:
                r, c = s['radius'], s['center']
                xs.extend([c[0] - r, c[0] + r])
                ys.extend([c[1] - r, c[1] + r])

        if not xs or not ys:
            messagebox.showwarning("警告", "描画可能なテキストや図形が見つかりませんでした。", parent=self)
            return

        self.min_x, max_x = min(xs), max(xs)
        min_y, self.max_y = min(ys), max(ys)
        
        w = max_x - self.min_x
        h = self.max_y - min_y
        if w == 0: w = 1
        if h == 0: h = 1
        
        self.base_scale = min(900 / w, 700 / h) * 0.9
        self.scale = self.base_scale
        self.offset_x = 50
        self.offset_y = 50
        
    def d2c(self, x, y):
        cx = (x - self.min_x) * self.scale + self.offset_x
        cy = (self.max_y - y) * self.scale + self.offset_y
        return cx, cy
        
    def c2d(self, cx, cy):
        x = (cx - self.offset_x) / self.scale + self.min_x
        y = self.max_y - (cy - self.offset_y) / self.scale
        return x, y
        
    def get_extracted_text(self):
        if not self.anchor or not self.rect_dxf_start or not self.rect_dxf_end:
            return ""
            
        x1, x2 = min(self.rect_dxf_start[0], self.rect_dxf_end[0]), max(self.rect_dxf_start[0], self.rect_dxf_end[0])
        y1, y2 = min(self.rect_dxf_start[1], self.rect_dxf_end[1]), max(self.rect_dxf_start[1], self.rect_dxf_end[1])
        
        ax, ay = self.anchor["x"], self.anchor["y"]
        xmin, xmax = x1 - ax, x2 - ax
        ymin, ymax = y1 - ay, y2 - ay
        
        matched_texts = []
        for t in self.texts:
            if t == self.anchor or t['text'].strip() == self.anchor['text'].strip(): continue
            dx = t['x'] - ax
            dy = t['y'] - ay
            if xmin <= dx <= xmax and ymin <= dy <= ymax:
                matched_texts.append(t)
        
        if matched_texts:
            matched_texts.sort(key=lambda item: (-item['y'], item['x']))
            return " ".join([t['text'] for t in matched_texts])
        return ""

    def update_preview_text(self):
        if not self.anchor:
            self.preview_label.config(text="【抽出プレビュー】 1. まずは図面内の基準にする文字をクリックしてください")
        elif not self.rect_dxf_start or not self.rect_dxf_end:
            self.preview_label.config(text=f"【基準文字: {self.anchor['text']}】 2. 次に抽出したい値の範囲を左ドラッグで囲んでください")
        else:
            ext_text = self.get_extracted_text()
            if ext_text:
                self.preview_label.config(text=f"【基準文字: {self.anchor['text']} ➔ 抽出テスト】 {ext_text}")
            else:
                self.preview_label.config(text=f"【基準文字: {self.anchor['text']}】 (※選択範囲内にテキストが見つかりません)")

    def draw(self, update_text=True):
        self.canvas.delete("all")
        cw = max(10, self.canvas.winfo_width())
        ch = max(10, self.canvas.winfo_height())
        margin = 100
        shape_color = "#C0C0C0"
        
        for s in self.shapes:
            try:
                if s['type'] == 'line':
                    cx1, cy1 = self.d2c(*s['start'])
                    cx2, cy2 = self.d2c(*s['end'])
                    if max(cx1, cx2) < -margin or min(cx1, cx2) > cw + margin or max(cy1, cy2) < -margin or min(cy1, cy2) > ch + margin:
                        continue
                    self.canvas.create_line(cx1, cy1, cx2, cy2, fill=shape_color)
                elif s['type'] == 'polyline':
                    pts = s['points']
                    if not pts: continue
                    c_pts = [self.d2c(*pt) for pt in pts]
                    if s.get('closed') and len(c_pts) > 2: c_pts.append(c_pts[0])
                    
                    cxs = [pt[0] for pt in c_pts]; cys = [pt[1] for pt in c_pts]
                    if max(cxs) < -margin or min(cxs) > cw + margin or max(cys) < -margin or min(cys) > ch + margin: continue
                    flat_pts = [coord for pt in c_pts for coord in pt]
                    self.canvas.create_line(flat_pts, fill=shape_color)
                elif s['type'] == 'circle':
                    cx, cy = self.d2c(*s['center'])
                    cr = s['radius'] * self.scale
                    if cx + cr < -margin or cx - cr > cw + margin or cy + cr < -margin or cy - cr > ch + margin: continue
                    self.canvas.create_oval(cx - cr, cy - cr, cx + cr, cy + cr, outline=shape_color)
                elif s['type'] == 'arc':
                    cx, cy = self.d2c(*s['center'])
                    cr = s['radius'] * self.scale
                    if cx + cr < -margin or cx - cr > cw + margin or cy + cr < -margin or cy - cr > ch + margin: continue
                    tk_extent = s['end_angle'] - s['start_angle']
                    if tk_extent < 0: tk_extent += 360
                    tk_start = 360 - s['end_angle']
                    self.canvas.create_arc(cx - cr, cy - cr, cx + cr, cy + cr, start=tk_start, extent=tk_extent, outline=shape_color, style=ARC)
            except: pass

        for t in self.texts:
            cx, cy = self.d2c(t["x"], t["y"])
            if cx < -margin or cx > cw + margin or cy < -margin or cy > ch + margin: continue
                
            color = "black"
            px_height = int(t.get("h", 2.5) * self.scale)
            f_size = max(2, px_height)
            
            if self.anchor and self.anchor == t:
                color = "red"
                f_size = int(f_size * 1.5)
                self.canvas.create_oval(cx-5, cy-5, cx+5, cy+5, fill="red", outline="red")
            self.canvas.create_text(cx, cy, text=t["text"], anchor="sw", fill=color, font=("Meiryo UI", -f_size))
            
        if self.rect_dxf_start and self.rect_dxf_end:
            cx1, cy1 = self.d2c(*self.rect_dxf_start)
            cx2, cy2 = self.d2c(*self.rect_dxf_end)
            self.canvas.create_rectangle(cx1, cy1, cx2, cy2, outline="blue", dash=(4, 4), width=2, fill="#e6f2ff", stipple="gray25")
            
        if update_text:
            self.update_preview_text()
            
    def on_left_press(self, event):
        dx, dy = self.c2d(event.x, event.y)
        if self.mode.get() == "anchor":
            best_t, best_dist = None, float('inf')
            for t in self.texts:
                dist = (t["x"] - dx)**2 + (t["y"] - dy)**2
                if dist < best_dist:
                    best_dist = dist; best_t = t
            if best_t:
                self.anchor = best_t
                self.mode.set("rect")
                self.draw()
        elif self.mode.get() == "rect":
            if not self.anchor:
                messagebox.showinfo("案内", "先に「1. 基準文字」を選択してください。", parent=self)
                self.mode.set("anchor")
                return
            self.rect_dxf_start = (dx, dy)
            self.rect_dxf_end = (dx, dy)
            self.is_dragging = True
            
    def on_left_drag(self, event):
        if self.mode.get() == "rect" and self.is_dragging:
            dx, dy = self.c2d(event.x, event.y)
            self.rect_dxf_end = (dx, dy)
            self.draw()
            
    def on_left_release(self, event):
        if self.mode.get() == "rect" and self.is_dragging:
            self.is_dragging = False
            self.draw()
            
    def on_right_press(self, event):
        self.pan_start_x, self.pan_start_y = event.x, event.y
        
    def on_right_drag(self, event):
        self.offset_x += (event.x - self.pan_start_x)
        self.offset_y += (event.y - self.pan_start_y)
        self.pan_start_x, self.pan_start_y = event.x, event.y
        self.draw()
        
    def on_mousewheel(self, event):
        zoom = 1.2 if event.delta > 0 else 0.8
        mx, my = event.x, event.y
        dx, dy = self.c2d(mx, my)
        self.scale *= zoom
        self.offset_x = mx - (dx - self.min_x) * self.scale
        self.offset_y = my - (self.max_y - dy) * self.scale
        self.draw()
        
    def confirm(self):
        if not self.anchor:
            messagebox.showwarning("エラー", "基準文字が選択されていません。", parent=self)
            return
        if not self.rect_dxf_start or not self.rect_dxf_end:
            messagebox.showwarning("エラー", "抽出範囲が選択されていません。\n左ドラッグで範囲を四角く囲んでください。", parent=self)
            return
            
        x1, x2 = min(self.rect_dxf_start[0], self.rect_dxf_end[0]), max(self.rect_dxf_start[0], self.rect_dxf_end[0])
        y1, y2 = min(self.rect_dxf_start[1], self.rect_dxf_end[1]), max(self.rect_dxf_start[1], self.rect_dxf_end[1])
        
        ax, ay = self.anchor["x"], self.anchor["y"]
        xmin, xmax = x1 - ax, x2 - ax
        ymin, ymax = y1 - ay, y2 - ay
        
        ext_text = self.get_extracted_text()
        kw_text = self.anchor["text"]
        
        self.on_complete(kw_text, "左下", xmin, xmax, ymin, ymax, ext_text)
        
        # 連続追加のためのリセット
        self.anchor = None
        self.rect_dxf_start = None
        self.rect_dxf_end = None
        self.mode.set("anchor")
        
        self.draw(update_text=False)
        self.preview_label.config(text=f"✅ 「{kw_text}」を追加しました！続けて次の基準文字をクリックしてください。")

def open_preview():
    target_files = [f for f in selected_files if f.lower().endswith('.dxf')]
    if current_mode == "folder" and selected_folder:
        target_files = [os.path.join(selected_folder, f) for f in os.listdir(selected_folder) if f.lower().endswith('.dxf')]
        
    if not target_files:
        messagebox.showwarning("警告", "対象内にDXFファイルが見つからないためプレビューを開けません。")
        return
        
    PreviewDialog(root, target_files[0], add_keyword_row)


# ==========================
# メイン抽出処理
# ==========================
def extract_dxf_text():
    try:
        target_files = []
        if current_mode == "file":
            target_files = [f for f in selected_files if f.lower().endswith('.dxf')]
        elif current_mode == "folder" and selected_folder:
            target_files = [os.path.join(selected_folder, f) for f in os.listdir(selected_folder) if f.lower().endswith('.dxf')]

        if not target_files:
            messagebox.showwarning("警告", "対象のDXFファイルがありません。")
            return

        save_dir = get_save_dir(target_files[0])
        if not save_dir: return 

        is_keyword_mode = (mode_var.get() == 1)
        y_threshold = threshold_var.get()
        processed_count = 0
        error_logs = []

        if is_keyword_mode:
            if not keyword_entries:
                messagebox.showwarning("警告", "抽出設定がありません。「＋ プレビューを開いて抽出範囲を設定」から設定を行ってください。")
                return

            out_wb = openpyxl.Workbook()
            ws = out_wb.active
            ws.title = "プレビュー抽出結果"
            
            ws_coord = out_wb.create_sheet(title="座標リスト(設定確認用)")
            ws_coord.append(["元ファイル名", "テキスト", "X座標", "Y座標"])
            ws_coord.column_dimensions['A'].width = 25
            ws_coord.column_dimensions['B'].width = 40
            ws_coord.column_dimensions['C'].width = 15
            ws_coord.column_dimensions['D'].width = 15
            
            header = ["元ファイル名"] + [cfg["kw_var"].get().strip() for cfg in keyword_entries]
            ws.append(header)

            for file_path in target_files:
                texts, _ = get_all_elements_from_dxf(file_path)
                if not texts:
                    error_logs.append(f"【{os.path.basename(file_path)}】テキストが見つかりません。")
                    continue

                for t in texts:
                    ws_coord.append([os.path.basename(file_path), t['text'], round(t['x'], 3), round(t['y'], 3)])

                file_row = [os.path.basename(file_path)]

                for kw_cfg in keyword_entries:
                    kw_clean = kw_cfg["kw_var"].get().replace(" ", "").replace("　", "").lower()
                    try:
                        x_min, x_max = float(kw_cfg["xmin_var"].get()), float(kw_cfg["xmax_var"].get())
                        y_min, y_max = float(kw_cfg["ymin_var"].get()), float(kw_cfg["ymax_var"].get())
                    except:
                        x_min, x_max, y_min, y_max = 0.0, 0.0, 0.0, 0.0
                        
                    found_val = ""
                    kw_entity = next((t for t in texts if kw_clean == t['text'].replace(" ", "").replace("　", "").lower()), None)
                    if not kw_entity:
                        kw_entity = next((t for t in texts if kw_clean in t['text'].replace(" ", "").replace("　", "").lower()), None)
                    
                    if kw_entity:
                        kx, ky = kw_entity['x'], kw_entity['y']
                        matched_texts = []
                        for t in texts:
                            if t == kw_entity or t['text'].strip() == kw_entity['text'].strip(): continue
                            dx, dy = t['x'] - kx, t['y'] - ky
                            if x_min <= dx <= x_max and y_min <= dy <= y_max:
                                matched_texts.append(t)
                        
                        if matched_texts:
                            matched_texts.sort(key=lambda item: (-item['y'], item['x']))
                            found_val = " ".join([t['text'] for t in matched_texts])
                    file_row.append(found_val)
                
                ws.append(file_row)
                processed_count += 1

            for col in ws.columns:
                try:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            length = sum(2 if ord(c) > 255 else 1 for c in str(cell.value))
                            if length > max_length: max_length = length
                    ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
                except: pass

            import time
            output_path = os.path.join(save_dir, f"DXFプレビュー抽出_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
            try: out_wb.save(output_path)
            except PermissionError:
                messagebox.showerror("保存エラー", f"Excelファイルが開かれているため保存できません。\nパス: {output_path}")
                return
            except Exception as e:
                messagebox.showerror("保存エラー", f"エラー: {e}")
                return
            
            msg = f"{processed_count} 個のファイルから抽出を完了しました。\n保存先: {output_path}"
            if error_logs: msg += "\n\n※一部エラーあり:\n" + "\n".join(error_logs[:5])
            messagebox.showinfo("完了", msg)

        else:
            # ---------------------------------------------------------
            # モードB: 全体抽出モード
            # ---------------------------------------------------------
            for file_path in target_files:
                texts, _ = get_all_elements_from_dxf(file_path)
                if not texts:
                    error_logs.append(f"【{os.path.basename(file_path)}】読込エラー。")
                    continue

                texts.sort(key=lambda item: (-item['y'], item['x']))
                rows, current_row, current_y = [], [], None

                for item in texts:
                    if current_y is None:
                        current_y = item['y']
                        current_row.append(item)
                    elif abs(current_y - item['y']) <= y_threshold:
                        current_row.append(item)
                    else:
                        current_row.sort(key=lambda i: i['x'])
                        rows.append([i['text'] for i in current_row])
                        current_y = item['y']
                        current_row = [item]
                if current_row:
                    current_row.sort(key=lambda i: i['x'])
                    rows.append([i['text'] for i in current_row])

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "抽出データ"
                for row_data in rows: ws.append(row_data)

                for col in ws.columns:
                    try:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            if cell.value:
                                length = sum(2 if ord(c) > 255 else 1 for c in str(cell.value))
                                if length > max_length: max_length = length
                        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
                    except: pass

                ws_coord = wb.create_sheet(title="座標リスト(設定確認用)")
                ws_coord.append(["テキスト", "X座標", "Y座標"])
                for item in texts: ws_coord.append([item['text'], round(item['x'], 3), round(item['y'], 3)])
                ws_coord.column_dimensions['A'].width, ws_coord.column_dimensions['B'].width, ws_coord.column_dimensions['C'].width = 40, 15, 15

                output_path = os.path.join(save_dir, f"{os.path.splitext(os.path.basename(file_path))[0]}_テキスト抽出.xlsx")
                try: wb.save(output_path)
                except PermissionError:
                    messagebox.showerror("保存エラー", f"Excelが開かれています。\nパス: {output_path}")
                    return
                except: return
                processed_count += 1

            msg = f"{processed_count} 個のDXFファイルからの全体抽出が完了しました。"
            if error_logs: msg += "\n\n※一部エラーあり:\n" + "\n".join(error_logs[:5])
            messagebox.showinfo("完了", msg)

    except Exception as e:
        messagebox.showerror("抽出エラー", f"予期せぬエラーが発生しました。\n\n【詳細】\n{traceback.format_exc()}")

# ==========================
# データ集約処理
# ==========================
def aggregate_data():
    try:
        target_files = []
        if current_mode == "file":
            target_files = [f for f in selected_files if f.lower().endswith(('.xlsx', '.xlsm', '.xls'))]
        elif current_mode == "folder" and selected_folder:
            target_files = [os.path.join(selected_folder, f) for f in os.listdir(selected_folder) if f.lower().endswith(('.xlsx', '.xlsm', '.xls')) and "集約" not in f]
            
        if not target_files:
            messagebox.showwarning("警告", "選択された対象内に集約可能なExcelファイルが見つかりません。")
            return

        save_dir = get_save_dir(target_files[0])
        if not save_dir: return

        agg_header = ["元ファイル名"]
        agg_rows, error_logs = [], []

        for f in target_files:
            fname = os.path.basename(f)
            try:
                wb = openpyxl.load_workbook(f, data_only=True)
                sheet = wb.worksheets[0]
                rows = list(sheet.iter_rows(values_only=True))
                wb.close()

                if not rows: continue
                valid_rows = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
                if not valid_rows: continue

                curr_header, curr_data = valid_rows[0], valid_rows[1:]
                safe_header = [str(h).strip() if h is not None and str(h).strip() else f"列{i+1}" for i, h in enumerate(curr_header)]
                col_mapping = {}

                for i, h in enumerate(safe_header):
                    if h not in agg_header: agg_header.append(h)
                    col_mapping[i] = agg_header.index(h)

                for r in curr_data:
                    row = [""] * len(agg_header)
                    row[0] = fname
                    for i, val in enumerate(r):
                        if i in col_mapping:
                            idx = col_mapping[i]
                            if idx >= len(row): row.extend([""] * (idx - len(row) + 1))
                            row[idx] = "" if val is None or str(val).strip() == "None" else str(val).strip()
                    agg_rows.append(row)
            except Exception as e:
                error_logs.append(f"【{fname}】スキップ: {e}")

        if not agg_rows:
            messagebox.showinfo("結果", "集約できるデータがありませんでした。")
            return

        out_wb = openpyxl.Workbook()
        ws = out_wb.active
        ws.title = "集約"

        final_data = [agg_header] + [r + [""] * (len(agg_header) - len(r)) for r in agg_rows]
        apply_text_inheritance(final_data)

        for row_data in final_data: ws.append(row_data)

        for col in ws.columns:
            try:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        length = sum(2 if ord(c) > 255 else 1 for c in str(cell.value))
                        if length > max_length: max_length = length
                ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
            except: pass

        import time
        output_path = os.path.join(save_dir, f"データ集約_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        try: out_wb.save(output_path)
        except PermissionError:
            messagebox.showerror("保存エラー", f"Excelファイルが開かれているため保存できません。\nパス: {output_path}")
            return
        except Exception as e:
            messagebox.showerror("保存エラー", f"エラー: {e}")
            return

        msg = f"{len(target_files)}個のファイルデータを集約しました。\n保存先: {output_path}"
        if error_logs: msg += f"\n\n※一部スキップ:\n" + "\n".join(error_logs[:3])
        messagebox.showinfo("完了", msg)

    except Exception as e:
        messagebox.showerror("集約エラー", f"エラーが発生しました。\n\n【詳細】\n{traceback.format_exc()}")

# ==========================
# UI構築
# ==========================
root = Tk()
root.title(f"{APP_TITLE} {VERSION}")
root.geometry("860x850")
root.minsize(800, 600)
root.configure(bg="#F8F9FA")

style = ttk.Style()
style.theme_use("clam")

header_frame = Frame(root, bg="#0D6EFD", pady=15)
header_frame.pack(fill=X)
Label(header_frame, text=f"{APP_TITLE} {VERSION}", font=("Meiryo UI", 16, "bold"), bg="#0D6EFD", fg="white").pack()
Label(header_frame, text="DXFテキスト抽出 ＆ Excelスマート集約ツール", font=("Meiryo UI", 10), bg="#0D6EFD", fg="white").pack()

container = Frame(root, bg="#F8F9FA")
container.pack(fill=BOTH, expand=True)
canvas = Canvas(container, bg="#F8F9FA", highlightthickness=0)
scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
scrollable_frame = Frame(canvas, bg="#F8F9FA", padx=20, pady=10)

scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", width=canvas.winfo_width())
canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas.find_withtag("all")[0], width=e.width))
canvas.configure(yscrollcommand=scrollbar.set)
canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")
root.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

main_frame = scrollable_frame

# --- 1. 対象の選択 ---
Label(main_frame, text="1. 対象の選択", font=("Meiryo UI", 11, "bold"), bg="#F8F9FA").pack(anchor=W, pady=(10, 5))
btn_frame = Frame(main_frame, bg="#F8F9FA")
btn_frame.pack(fill=X)
Button(btn_frame, text="📄 ファイルを選択", command=select_files, width=20, bg="#E9ECEF", relief=GROOVE).pack(side=LEFT, padx=5, pady=5)
Button(btn_frame, text="📁 フォルダを選択", command=select_folder, width=20, bg="#E9ECEF", relief=GROOVE).pack(side=LEFT, padx=5, pady=5)

Label(main_frame, text="選択中のパス:", bg="#F8F9FA").pack(anchor=W, pady=(5, 0))
text_paths = Text(main_frame, height=3, font=("Meiryo UI", 9))
text_paths.pack(fill=X, pady=5)

# --- 2. 設定 ---
Label(main_frame, text="2. 設定", font=("Meiryo UI", 11, "bold"), bg="#F8F9FA").pack(anchor=W, pady=(15, 5))
settings_frame = Frame(main_frame, bg="#FFFFFF", padx=15, pady=15, relief=SOLID, bd=1)
settings_frame.pack(fill=X, pady=5)

Label(settings_frame, text="【保存先】", font=("Meiryo UI", 9, "bold"), bg="#FFFFFF").pack(anchor=W)
save_option = IntVar(value=1)
Radiobutton(settings_frame, text="元のファイルと同じフォルダに保存", variable=save_option, value=1, bg="#FFFFFF").pack(anchor=W)
Radiobutton(settings_frame, text="実行時に任意のフォルダを指定して保存", variable=save_option, value=2, bg="#FFFFFF").pack(anchor=W)

Frame(settings_frame, height=1, bg="#DEE2E6").pack(fill=X, pady=10)

Label(settings_frame, text="【DXF抽出 詳細設定】", font=("Meiryo UI", 9, "bold"), bg="#FFFFFF").pack(anchor=W)

# --- 保存・読込エリア ---
settings_btn_frame = Frame(settings_frame, bg="#FFFFFF")
settings_btn_frame.pack(fill=X, pady=(0, 10))
Button(settings_btn_frame, text="💾 現在の抽出設定を保存", command=save_settings, bg="#E9ECEF", relief=GROOVE).pack(side=RIGHT, padx=2)
Button(settings_btn_frame, text="📂 保存した抽出設定を読込", command=load_settings, bg="#E9ECEF", relief=GROOVE).pack(side=RIGHT, padx=2)
# ------------------------

mode_var = IntVar(value=1)
def toggle_mode():
    state = NORMAL if mode_var.get() == 1 else DISABLED
    for child in kw_container.winfo_children():
        try: child.configure(state=state)
        except: pass
    for row in keyword_ui_frames:
        for child in row.winfo_children():
            try: child.configure(state=state)
            except: pass

Radiobutton(settings_frame, text="1. プレビュー画面から抽出範囲を選択して抽出（名寄せ・集約用）", variable=mode_var, value=1, bg="#FFFFFF", command=toggle_mode).pack(anchor=W, pady=(5, 0))

kw_container = Frame(settings_frame, bg="#F0F4F8", padx=10, pady=10)
kw_container.pack(fill=X, pady=5, padx=20)

Label(kw_container, text="▼ 設定済みのキーワードと抽出範囲", font=("Meiryo UI", 9, "bold"), bg="#F0F4F8").pack(anchor=W)
Label(kw_container, text="※ プレビュー画面を開いたまま、複数の抽出項目を連続して設定できます。", font=("Meiryo UI", 8), fg="#6C757D", bg="#F0F4F8").pack(anchor=W, pady=(0, 5))

Button(kw_container, text="＋ プレビューを開いて抽出範囲を設定", command=open_preview, bg="#0D6EFD", fg="white", font=("Meiryo UI", 9, "bold"), padx=10).pack(anchor=W, pady=5)

kw_list_frame = Frame(kw_container, bg="#F0F4F8")
kw_list_frame.pack(fill=X, pady=5)

keyword_entries = []     # 内部データ用
keyword_ui_frames = []   # UI削除用

def add_keyword_row(kw="", anchor_val="左下", xmin=0.0, xmax=0.0, ymin=0.0, ymax=0.0, sample_text=""):
    row_frame = Frame(kw_list_frame, bg="#FFFFFF", pady=6, padx=10, relief=SOLID, bd=1)
    row_frame.pack(fill=X, pady=3)

    top_frame = Frame(row_frame, bg="#FFFFFF")
    top_frame.pack(fill=X, pady=(0, 3))
    
    Label(top_frame, text="基準文字:", font=("Meiryo UI", 8, "bold"), bg="#FFFFFF", fg="#495057").pack(side=LEFT)
    
    kw_var = StringVar(value=kw)
    Entry(top_frame, textvariable=kw_var, width=15).pack(side=LEFT, padx=5)

    anchors = ["左下", "右下", "左上", "右上", "中央"]
    anchor_var = StringVar(value=anchor_val)
    cb = ttk.Combobox(top_frame, textvariable=anchor_var, values=anchors, width=5, state="readonly")
    cb.pack(side=LEFT, padx=2)

    if sample_text:
        Label(top_frame, text=f"⇒ 抽出サンプル: {sample_text}", font=("Meiryo UI", 9, "bold"), bg="#FFFFFF", fg="#198754").pack(side=LEFT, padx=15)

    bottom_frame = Frame(row_frame, bg="#FFFFFF")
    bottom_frame.pack(fill=X)

    xmin_var = DoubleVar(value=xmin)
    xmax_var = DoubleVar(value=xmax)
    ymin_var = DoubleVar(value=ymin)
    ymax_var = DoubleVar(value=ymax)

    Label(bottom_frame, text="X範囲:", bg="#FFFFFF", fg="#0D6EFD").pack(side=LEFT)
    Spinbox(bottom_frame, textvariable=xmin_var, from_=-5000, to=5000, width=6, format="%.1f").pack(side=LEFT)
    Label(bottom_frame, text="~", bg="#FFFFFF").pack(side=LEFT)
    Spinbox(bottom_frame, textvariable=xmax_var, from_=-5000, to=5000, width=6, format="%.1f").pack(side=LEFT)

    Frame(bottom_frame, width=1, bg="#DEE2E6").pack(side=LEFT, fill=Y, padx=10)

    Label(bottom_frame, text="Y範囲:", bg="#FFFFFF", fg="#198754").pack(side=LEFT)
    Spinbox(bottom_frame, textvariable=ymin_var, from_=-5000, to=5000, width=6, format="%.1f").pack(side=LEFT)
    Label(bottom_frame, text="~", bg="#FFFFFF").pack(side=LEFT)
    Spinbox(bottom_frame, textvariable=ymax_var, from_=-5000, to=5000, width=6, format="%.1f").pack(side=LEFT)

    def remove_row():
        row_frame.destroy()
        for item in keyword_entries:
            if item["frame"] == row_frame:
                keyword_entries.remove(item)
                break
        if row_frame in keyword_ui_frames:
            keyword_ui_frames.remove(row_frame)
        root.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))

    Button(top_frame, text="削除", command=remove_row, bg="#DC3545", fg="white", padx=10).pack(side=RIGHT, padx=5)

    row_data = {
        "kw_var": kw_var, "anchor_var": anchor_var, "xmin_var": xmin_var, "xmax_var": xmax_var,
        "ymin_var": ymin_var, "ymax_var": ymax_var, "frame": row_frame, "sample_text": sample_text
    }
    keyword_entries.append(row_data)
    keyword_ui_frames.append(row_frame)
    
    root.update_idletasks()
    canvas.configure(scrollregion=canvas.bbox("all"))

Radiobutton(settings_frame, text="2. ファイルごとに図面内のすべてのテキストを抽出", variable=mode_var, value=2, bg="#FFFFFF", command=toggle_mode).pack(anchor=W, pady=(15, 0))
Label(settings_frame, text="※全体抽出モード用のテキストズレ許容値（閾値）:", bg="#FFFFFF").pack(anchor=W, pady=(5, 0))
threshold_var = DoubleVar(value=20.0)
Spinbox(settings_frame, from_=0.0, to=500.0, increment=1.0, textvariable=threshold_var, width=10).pack(anchor=W, pady=2)


# --- 3. 実行 ---
Label(main_frame, text="3. 実行", font=("Meiryo UI", 11, "bold"), bg="#F8F9FA").pack(anchor=W, pady=(20, 5))
action_frame = Frame(main_frame, bg="#F8F9FA")
action_frame.pack(fill=X)

btn_extract = Button(action_frame, text="🚀 選択対象のDXFからテキストを抽出", command=extract_dxf_text, height=2, bg="#198754", fg="white", font=("Meiryo UI", 10, "bold"), state=DISABLED)
btn_extract.pack(fill=X, pady=5)

btn_aggregate = Button(action_frame, text="🧩 選択対象のExcelを1つにスマート集約", command=aggregate_data, height=2, bg="#6F42C1", fg="white", font=("Meiryo UI", 10, "bold"), state=DISABLED)
btn_aggregate.pack(fill=X, pady=5)


if __name__ == "__main__":
    root.mainloop()